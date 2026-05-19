import csv
import random
import os
import calendar
from collections import defaultdict
from datetime import date, timedelta, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
HARD_START_DATE = "2026-07-01"
FRI_SUN_PAIRING_BONUS = -50 

class Resident:
    def __init__(self, name, level, block_units, base_priority, weight):
        self.name = name
        self.level = level.lower()
        self.block_budgets = block_units 
        self.base_priority = base_priority 
        self.weight = float(weight)
        self.h1_units = 0
        self.h2_units = 0
        self.total_assigned_units = 0
        self.total_shifts = 0
        self.block_assignments = defaultdict(int)
        self.active_priorities = {}        # {wk_idx: {"level": int, "type": "hard"|"soft"}}
        self._priority_dates = {}          
        self._assignments = []

    def remaining_in_block(self, block_num):
        return self.block_budgets.get(block_num, 0) - self.block_assignments[block_num]

class CallScheduler:
    def __init__(self, residents, n_blocks=13):
        self.residents = residents
        self.n_blocks = n_blocks
        self.n_weekends = n_blocks * 4
        base_date = date.fromisoformat(HARD_START_DATE)
        days_to_friday = (4 - base_date.weekday() + 7) % 7
        self.start_friday = base_date + timedelta(days=days_to_friday)
        self.schedule = [
            {d: {"senior": None, "junior": None} for d in ["friday", "saturday", "sunday"]} 
            for _ in range(self.n_weekends)
        ]
        self.violations = [] 

    def _can_assign(self, res, wk, day):
        block_num = (wk // 4) + 1
        u = {"friday": 1, "saturday": 2, "sunday": 1}[day]
        if res.remaining_in_block(block_num) < u: return False
        if self.schedule[wk][day][res.level]: return False
        
        wk_units = sum({"friday": 1, "saturday": 2, "sunday": 1}[d] for w, d in res._assignments if w == wk)
        if wk_units + u > 2: return False
        
        days_on = [d for w, d in res._assignments if w == wk]
        if day == "saturday" and ("friday" in days_on or "sunday" in days_on): return False
        if (day == "friday" or day == "sunday") and "saturday" in days_on: return False
        return True

    def _apply_fixed_assignment(self, res_name, target_date, role):
        delta = target_date - self.start_friday
        wk_idx = delta.days // 7
        day_map = {0: "friday", 1: "saturday", 2: "sunday"}
        day_str = day_map.get(delta.days % 7)
        if 0 <= wk_idx < self.n_weekends and day_str:
            res = next((r for r in self.residents if r.name == res_name), None)
            if res:
                u = {"friday": 1, "saturday": 2, "sunday": 1}[day_str]
                self.schedule[wk_idx][day_str][role] = res.name
                block_num = (wk_idx // 4) + 1
                if block_num <= 7: res.h1_units += u
                else: res.h2_units += u
                res.block_assignments[block_num] += u
                res.total_assigned_units += u
                res.total_shifts += 1
                res._assignments.append((wk_idx, day_str))
                return True
        return False

    def build(self):
        for wk in range(self.n_weekends):
            for role in ["senior", "junior"]:
                for day in ["friday", "saturday", "sunday"]:
                    if self.schedule[wk][day][role]: continue
                    pri_fri = (day == "sunday")
                    self._fill_slot(wk, day, role, prioritize_friday_person=pri_fri)
        self._check_priority_violations()

    def _fill_slot(self, wk, day, role, prioritize_friday_person=False):
        pool = [r for r in self.residents if r.level == role and self._can_assign(r, wk, day)]
        if not pool: return
        fri_person = self.schedule[wk]["friday"][role]
        u = {"friday": 1, "saturday": 2, "sunday": 1}[day]
        
        group_avg_units = sum(r.total_assigned_units for r in pool) / len(pool) if pool else 0
        
        def score_func(res):
            req = res.active_priorities.get(wk, {})
            req_type = req.get("type", "none")
            
            prio_penalty = 0
            if req_type == "hard": prio_penalty = 20000000  
            elif req_type == "soft": prio_penalty = 500000    
            
            set_bonus = FRI_SUN_PAIRING_BONUS if prioritize_friday_person and fri_person == res.name else 0
            
            projected_units = res.total_assigned_units + u
            diff_from_avg = projected_units - group_avg_units
            
            work_load_score = (projected_units / res.weight) * 1000
            variance_penalty = (diff_from_avg ** 3) * 5000 if diff_from_avg > 0 else 0
            
            return prio_penalty + set_bonus + work_load_score + variance_penalty + random.uniform(0, 0.1)

        pool.sort(key=score_func)
        chosen = pool[0]
        self.schedule[wk][day][role] = chosen.name
        
        block_num = (wk // 4) + 1
        if block_num <= 7: chosen.h1_units += u
        else: chosen.h2_units += u
        chosen.block_assignments[block_num] += u
        chosen.total_assigned_units += u
        chosen.total_shifts += 1
        chosen._assignments.append((wk, day))

    def _check_priority_violations(self):
        for res in self.residents:
            for wk_idx, original_date in res._priority_dates.items():
                assigned_days = [d for w, d in res._assignments if w == wk_idx]
                if assigned_days:
                    req_type = res.active_priorities.get(wk_idx, {}).get("type", "soft")
                    self.violations.append({
                        "name": res.name,
                        "date": original_date,
                        "type": req_type.upper(),
                        "priority": res.active_priorities[wk_idx].get("level", 0),
                        "assigned": ", ".join([d.capitalize() for d in assigned_days])
                    })

    def print_violation_report(self):
        print(f"\n{'!'*20} UNACCOMMODATED REQUESTS REPORT {'!'*20}")
        if not self.violations:
            print("All priority requests (Hard and Soft) were successfully accommodated!")
        else:
            print("The following requested dates were assigned due to strict balancing requirements:")
            print(f"{'Name':<15} | {'Date':<12} | {'Type':<6} | {'Level':<5} | {'Assigned Shift'}")
            print("-" * 65)
            for v in sorted(self.violations, key=lambda x: (x['type'] != 'HARD', x['date'])):
                print(f"{v['name']:<15} | {v['date']:<12} | {v['type']:<6} | P-{v['priority']:<3} | {v['assigned']}")
        print("-" * 65)

    def print_distribution_report(self):
        print(f"\n{'='*20} CALL DISTRIBUTION REPORT {'='*20}")
        block_cols = [f"B{b}" for b in range(1, self.n_blocks + 1)]
        col_w = 5
        header = f"{'Name':<18} {'Lvl':<4} {'Shifts':>6} {'H1':>{col_w}} {'H2':>{col_w}} {'Tot U':>6}  " + \
                 "  ".join(f"{b:>{col_w}}" for b in block_cols)
        print(header)
        print("-" * len(header))

        for role in ["senior", "junior"]:
            group = [r for r in self.residents if r.level == role]
            print(f"\n-- {role.capitalize()}s --")
            for res in sorted(group, key=lambda r: r.name):
                block_str = "  ".join(f"{res.block_assignments.get(b, 0):>{col_w}}" for b in range(1, self.n_blocks + 1))
                row = (f"{res.name:<18} {res.level.capitalize()[:3]:<4} "
                       f"{res.total_shifts:>6} {res.h1_units:>{col_w}} {res.h2_units:>{col_w}} {res.total_assigned_units:>6}  "
                       f"{block_str}")
                print(row)

            block_totals = [sum(r.block_assignments.get(b, 0) for r in self.residents if r.level == role) for b in range(1, self.n_blocks + 1)]
            shifts_tot = sum(r.total_shifts for r in self.residents if r.level == role)
            print("-" * len(header))
            bt_str = "  ".join(f"{v:>{col_w}}" for v in block_totals)
            print(f"{role.capitalize() + ' Total':<18} {'':4} {shifts_tot:>6} {sum(block_totals[:7]):>{col_w}} {sum(block_totals[7:]):>{col_w}} {sum(block_totals):>6}  {bt_str}")
        print("=" * len(header))

    def print_calendar_view(self, start_str=None, end_str=None):
        try:
            s_date = date.fromisoformat(start_str) if start_str else self.start_friday
            max_end = self.start_friday + timedelta(weeks=self.n_weekends)
            e_date = date.fromisoformat(end_str) if end_str else max_end
        except ValueError:
            print("Invalid date format. Using full schedule range.")
            s_date = self.start_friday; e_date = self.start_friday + timedelta(weeks=self.n_weekends)

        print(f"\n{'='*30} TERMINAL CALENDAR VIEW {'='*30}")
        shift_lookup = {}
        for wk in range(self.n_weekends):
            fri_date = self.start_friday + timedelta(weeks=wk)
            shift_lookup[fri_date] = self.schedule[wk]["friday"]
            shift_lookup[fri_date + timedelta(days=1)] = self.schedule[wk]["saturday"]
            shift_lookup[fri_date + timedelta(days=2)] = self.schedule[wk]["sunday"]

        cal = calendar.TextCalendar(calendar.SUNDAY)
        curr_year, curr_month = s_date.year, s_date.month
        end_year, end_month = e_date.year, e_date.month

        while (curr_year, curr_month) <= (end_year, end_month):
            print(f"\n--- {calendar.month_name[curr_month]} {curr_year} ---")
            print(f"{'Sun':<15}{'Mon':<15}{'Tue':<15}{'Wed':<15}{'Thu':<15}{'Fri':<15}{'Sat':<15}")
            print("-" * 105)

            for week in cal.monthdatescalendar(curr_year, curr_month):
                date_row = "".join(f"{d.day:<15}" if d.month == curr_month else f"{'':<15}" for d in week)
                print(date_row)
                senior_row = ""
                for d in week:
                    if d.month == curr_month and d in shift_lookup:
                        snr = shift_lookup[d]['senior'] or 'OPEN'
                        senior_row += f"S:{snr[:10]:<13}"
                    else: senior_row += f"{'':<15}"
                print(senior_row)
                junior_row = ""
                for d in week:
                    if d.month == curr_month and d in shift_lookup:
                        jnr = shift_lookup[d]['junior'] or 'OPEN'
                        junior_row += f"J:{jnr[:10]:<13}"
                    else: junior_row += f"{'':<15}"
                print(junior_row)
                print("-" * 105)

            curr_month += 1
            if curr_month > 12: curr_month = 1; curr_year += 1

    def print_text_output(self, target_block=None):
        title = "FULL YEAR SCHEDULE" if target_block is None else f"BLOCK {target_block} SCHEDULE"
        print(f"\n{'='*20} {title} {'='*20}")
        print(f"{'Date':<12} | {'Day':<10} | {'Senior':<18} | {'Junior':<18}")
        print("-" * 65)
        for wk in range(self.n_weekends):
            if target_block and (wk // 4) + 1 != target_block: continue
            fri_date = self.start_friday + timedelta(weeks=wk)
            for i, day in enumerate(["friday", "saturday", "sunday"]):
                d = fri_date + timedelta(days=i)
                s = self.schedule[wk][day]["senior"] or "OPEN"
                j = self.schedule[wk][day]["junior"] or "OPEN"
                print(f"{d.strftime('%m/%d/%Y'):<12} | {day.capitalize():<10} | {s:<18} | {j:<18}")
            print("-" * 65)

    def export_to_excel(self, filename="call_schedule.xlsx"):
        wb = Workbook()

        CLR_SENIOR   = "BFDFFF"   
        CLR_JUNIOR   = "D5F5D5"   
        CLR_HEADER   = "2F4F8F"   
        CLR_SUBHDR   = "4472C4"   
        CLR_WEEKEND  = "F2F2F2"   
        CLR_OPEN     = "FFE4E1"   
        CLR_SAT      = "FFF3CD"   
        WHITE        = "FFFFFF"

        thin = Side(style="thin", color="AAAAAA")
        med  = Side(style="medium", color="555555")
        def thin_border(): return Border(left=thin, right=thin, top=thin, bottom=thin)
        def hdr_font(sz=11, bold=True, color="FFFFFF"): return Font(name="Arial", size=sz, bold=bold, color=color)
        def body_font(sz=10, bold=False, color="000000"): return Font(name="Arial", size=sz, bold=bold, color=color)
        def fill(hex_color): return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
        def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
        def left(): return Alignment(horizontal="left", vertical="center")

        # ── Sheet 1: Tabular Schedule ──
        ws = wb.active
        ws.title = "Schedule"
        ws.freeze_panes = "A3"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 13
        ws.column_dimensions["C"].width = 11
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20

        ws.merge_cells("A1:E1")
        ws["A1"] = "RESIDENT CALL SCHEDULE"
        ws["A1"].font = hdr_font(sz=13)
        ws["A1"].fill = fill(CLR_HEADER)
        ws["A1"].alignment = center()

        for col, label in enumerate(["Block", "Date", "Day", "Senior On Call", "Junior On Call"], 1):
            cell = ws.cell(row=2, column=col, value=label)
            cell.font = hdr_font(sz=10)
            cell.fill = fill(CLR_SUBHDR)
            cell.alignment = center()
            cell.border = thin_border()

        row = 3
        for wk in range(self.n_weekends):
            block_num = (wk // 4) + 1
            fri_date = self.start_friday + timedelta(weeks=wk)

            for i, day in enumerate(["friday", "saturday", "sunday"]):
                d = fri_date + timedelta(days=i)
                senior = self.schedule[wk][day]["senior"] or "OPEN"
                junior = self.schedule[wk][day]["junior"] or "OPEN"
                is_sat = (day == "saturday")
                row_bg = CLR_SAT if is_sat else (CLR_WEEKEND if wk % 2 == 0 else WHITE)

                vals = [f"B{block_num}", d.strftime("%m/%d/%Y"), day.capitalize(), senior, junior]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border()
                    cell.alignment = center() if col < 4 else left()
                    cell.font = body_font()

                    if col == 4: cell.fill = fill(CLR_OPEN if senior == "OPEN" else CLR_SENIOR)
                    elif col == 5: cell.fill = fill(CLR_OPEN if junior == "OPEN" else CLR_JUNIOR)
                    else: cell.fill = fill(row_bg)
                row += 1
            for col in range(1, 6): ws.cell(row=row, column=col).border = Border(top=med)

        # ── Sheet 2: Distribution Summary ──
        ws2 = wb.create_sheet("Distribution")
        ws2.freeze_panes = "A3"
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 9
        ws2.column_dimensions["C"].width = 12  
        ws2.column_dimensions["D"].width = 7
        ws2.column_dimensions["E"].width = 7
        ws2.column_dimensions["F"].width = 11  
        for b in range(1, self.n_blocks + 1):
            ws2.column_dimensions[get_column_letter(6 + b)].width = 6

        last_col = get_column_letter(6 + self.n_blocks)
        ws2.merge_cells(f"A1:{last_col}1")
        ws2["A1"] = "CALL DISTRIBUTION SUMMARY"
        ws2["A1"].font = hdr_font(sz=12)
        ws2["A1"].fill = fill(CLR_HEADER)
        ws2["A1"].alignment = center()

        headers = ["Name", "Level", "Total Shifts", "H1", "H2", "Total Units"] + [f"B{b}" for b in range(1, self.n_blocks + 1)]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=2, column=col, value=h)
            cell.font = hdr_font(sz=10)
            cell.fill = fill(CLR_SUBHDR)
            cell.alignment = center()
            cell.border = thin_border()

        seniors = sorted([r for r in self.residents if r.level == "senior"], key=lambda r: r.name)
        juniors = sorted([r for r in self.residents if r.level == "junior"], key=lambda r: r.name)

        data_row = 3
        for group_label, group, grp_clr in [("Seniors", seniors, CLR_SENIOR), ("Juniors", juniors, CLR_JUNIOR)]:
            ws2.merge_cells(f"A{data_row}:{last_col}{data_row}")
            lbl_cell = ws2[f"A{data_row}"]
            lbl_cell.value = group_label
            lbl_cell.font = hdr_font(sz=10, color="000000")
            lbl_cell.fill = fill("DDDDDD")
            lbl_cell.alignment = left()
            data_row += 1

            for res in group:
                row_vals = [res.name, res.level.capitalize(), res.total_shifts, res.h1_units, res.h2_units, res.total_assigned_units] + \
                           [res.block_assignments.get(b, 0) for b in range(1, self.n_blocks + 1)]
                for col, val in enumerate(row_vals, 1):
                    cell = ws2.cell(row=data_row, column=col, value=val)
                    cell.font = body_font()
                    cell.alignment = center() if col > 1 else left()
                    cell.border = thin_border()
                    cell.fill = fill(grp_clr)
                    if col in [3, 6] and isinstance(val, int): cell.font = body_font(bold=True)
                data_row += 1

            role = group[0].level if group else ""
            totals = ["TOTAL", role.capitalize(), sum(r.total_shifts for r in group),
                      sum(r.h1_units for r in group), sum(r.h2_units for r in group),
                      sum(r.total_assigned_units for r in group)] + \
                     [sum(r.block_assignments.get(b, 0) for r in group) for b in range(1, self.n_blocks + 1)]
            for col, val in enumerate(totals, 1):
                cell = ws2.cell(row=data_row, column=col, value=val)
                cell.font = body_font(bold=True)
                cell.alignment = center() if col > 1 else left()
                cell.border = thin_border()
                cell.fill = fill("CCCCCC")
            data_row += 2 

        # ── Sheet 3: Visual Calendar Grid ──
        ws3 = wb.create_sheet("Visual Calendar")
        ws3.views.sheetView[0].showGridLines = True
        days_of_week = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        for col_idx, day_name in enumerate(days_of_week, 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 24

        shift_lookup = {}
        for wk in range(self.n_weekends):
            fri = self.start_friday + timedelta(weeks=wk)
            shift_lookup[fri] = self.schedule[wk]["friday"]
            shift_lookup[fri + timedelta(days=1)] = self.schedule[wk]["saturday"]
            shift_lookup[fri + timedelta(days=2)] = self.schedule[wk]["sunday"]

        end_date = self.start_friday + timedelta(weeks=self.n_weekends)
        current = date(self.start_friday.year, self.start_friday.month, 1)
        cal_engine = calendar.Calendar(calendar.SUNDAY)
        
        xl_row = 1
        while current <= end_date:
            ws3.merge_cells(start_row=xl_row, start_column=1, end_row=xl_row, end_column=7)
            m_cell = ws3.cell(row=xl_row, column=1, value=f"{calendar.month_name[current.month]} {current.year}")
            m_cell.font = hdr_font(sz=14, bold=True)
            m_cell.fill = fill(CLR_HEADER)
            m_cell.alignment = center()
            ws3.row_dimensions[xl_row].height = 28
            xl_row += 1
            
            for col_idx, day_name in enumerate(days_of_week, 1):
                c = ws3.cell(row=xl_row, column=col_idx, value=day_name)
                c.font = hdr_font(sz=10, bold=True)
                c.fill = fill(CLR_SUBHDR)
                c.alignment = center()
                c.border = thin_border()
            ws3.row_dimensions[xl_row].height = 20
            xl_row += 1
            
            for week in cal_engine.monthdatescalendar(current.year, current.month):
                ws3.row_dimensions[xl_row].height = 18
                for col_idx, d in enumerate(week, 1):
                    c = ws3.cell(row=xl_row, column=col_idx)
                    if d.month == current.month:
                        c.value = d.day
                        c.font = body_font(sz=10, bold=True)
                    c.fill = fill("F9F9F9" if d.month == current.month else "EAEAEA")
                    c.border = Border(left=thin, right=thin, top=thin)
                    c.alignment = Alignment(horizontal="left", vertical="top")
                xl_row += 1
                
                ws3.row_dimensions[xl_row].height = 20
                for col_idx, d in enumerate(week, 1):
                    c = ws3.cell(row=xl_row, column=col_idx)
                    if d.month == current.month and d in shift_lookup:
                        val = shift_lookup[d]["senior"]
                        c.value = f"S: {val}" if val else "S: OPEN"
                        c.fill = fill(CLR_SENIOR if val else CLR_OPEN)
                    else: c.fill = fill("F9F9F9" if d.month == current.month else "EAEAEA")
                    c.font = body_font(sz=9)
                    c.border = Border(left=thin, right=thin)
                xl_row += 1
                
                ws3.row_dimensions[xl_row].height = 20
                for col_idx, d in enumerate(week, 1):
                    c = ws3.cell(row=xl_row, column=col_idx)
                    if d.month == current.month and d in shift_lookup:
                        val = shift_lookup[d]["junior"]
                        c.value = f"J: {val}" if val else "J: OPEN"
                        c.fill = fill(CLR_JUNIOR if val else CLR_OPEN)
                    else: c.fill = fill("F9F9F9" if d.month == current.month else "EAEAEA")
                    c.font = body_font(sz=9)
                    c.border = Border(left=thin, right=thin, bottom=thin)
                xl_row += 2 
                
            current = date(current.year + 1, 1, 1) if current.month == 12 else date(current.year, current.month + 1, 1)
            xl_row += 1 

        wb.save(filename)
        print(f"\n✔ Schedule exported to: {filename}")
        return filename

def load_all_data():
    residents_map = {}
    try:
        with open('residents.csv', mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                budgets = {b: int(row.get(f'B{b}', 0)) for b in range(1, 14)}
                res = Resident(row['Name'], row['Level'], budgets, int(row.get('PriorityLevel', 0)), row.get('Weight', 1.0))
                residents_map[res.name] = res
    except: return None, None

    start_date = date.fromisoformat(HARD_START_DATE)
    first_friday = start_date + timedelta(days=(4 - start_date.weekday() + 7) % 7)

    try:
        with open('priorities.csv', mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row['Name']
                if name in residents_map:
                    date_str = row['Date'].strip()
                    p_type = row.get('Type', 'soft').strip().lower() 
                    if p_type not in ['hard', 'soft']: p_type = 'soft'
                    
                    wk_idx = (date.fromisoformat(date_str) - first_friday).days // 7
                    if 0 <= wk_idx < 52: 
                        residents_map[name].active_priorities[wk_idx] = {
                            "level": residents_map[name].base_priority,
                            "type": p_type
                        }
                        residents_map[name]._priority_dates[wk_idx] = date_str 
    except Exception as e: pass
    return list(residents_map.values()), first_friday

def load_hard_coded_excel(scheduler, filename="hard_coded.xlsx"):
    """Reads hard-coded assignments from an Excel file and applies them."""
    if not os.path.exists(filename): return
    try:
        wb = load_workbook(filename, data_only=True)
        ws = wb.active
        print(f"\nLoading hard-coded shifts from {filename}...")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, date_val, role = row[0], row[1], row[2]
            if name and date_val and role:
                d_str = date_val.date() if isinstance(date_val, datetime) else date.fromisoformat(str(date_val).split(' ')[0].strip())
                if scheduler._apply_fixed_assignment(name, d_str, role.lower()):
                    count += 1
        if count > 0: print(f"  ✔ Successfully locked {count} shifts from Excel.")
    except Exception as e: print(f"\nNotice: Error reading {filename} ({e})")

def load_incomplete_schedule(scheduler, filename="incomplete_schedule.csv"):
    """Safely reads a partially completed schedule from CSV."""
    if not os.path.exists(filename): return
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            shifts_count = 0
            for row in reader:
                if row.get('Name') and row.get('Date') and row.get('Role'):
                    if scheduler._apply_fixed_assignment(row['Name'].strip(), date.fromisoformat(row['Date'].strip()), row['Role'].strip().lower()):
                        shifts_count += 1
            if shifts_count > 0:
                print(f"\n✔ Seeded {shifts_count} historical shifts from {filename}. Re-balancing the rest...")
    except Exception as e: print(f"\nNotice: {filename} could not be processed ({e}). Generating schedule from baseline.")

if __name__ == "__main__":
    res_list, first_fri = load_all_data()
    if res_list:
        scheduler = CallScheduler(res_list)
        
        # 1. Apply legacy Fixed Schedule if exists
        try:
            with open('fixed_schedule.csv', mode='r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    scheduler._apply_fixed_assignment(row['Name'], date.fromisoformat(row['Date']), row['Role'].lower())
        except: pass

        # 2. Load Hard Coded Excel shifts
        load_hard_coded_excel(scheduler, "hard_coded.xlsx")

        # 3. Evaluate incomplete_schedule.csv (safe bypass)
        load_incomplete_schedule(scheduler, "incomplete_schedule.csv")

        # 4. Build remaining schedule around all loaded data
        scheduler.build()
        
        scheduler.print_violation_report()
        scheduler.print_distribution_report()
        scheduler.export_to_excel("call_schedule.xlsx")

        while True:
            print("\n--- VIEW OPTIONS ---")
            print("1. View List: Full Year")
            print("2. View List: Specific Block")
            print("3. View Distribution Report")
            print("4. View Monthly Calendar (Terminal Grid)")
            print("5. Export to Excel File")
            print("6. Exit")
            choice = input("Select (1/2/3/4/5/6): ")
            
            if choice == "1": scheduler.print_text_output()
            elif choice == "2":
                try:
                    b_num = int(input("Enter Block # (1-13): "))
                    if 1 <= b_num <= 13: scheduler.print_text_output(target_block=b_num)
                except: pass
            elif choice == "3": scheduler.print_distribution_report()
            elif choice == "4": 
                print("\nPress Enter to view the full year, or enter range bounds below:")
                s_input = input("Start Date (YYYY-MM-DD) or Enter: ").strip()
                e_input = input("End Date (YYYY-MM-DD) or Enter: ").strip()
                scheduler.print_calendar_view(start_str=s_input or None, end_str=e_input or None)
            elif choice == "5":
                fname = input("Filename (default: call_schedule.xlsx): ").strip() or "call_schedule.xlsx"
                scheduler.export_to_excel(fname)
            elif choice == "6": break
